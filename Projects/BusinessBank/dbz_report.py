"""
dbz_report.py — Полный финансовый отчёт ДБЗ (ИП Пирожкова Н.В.)

Собирает данные из всех источников и формирует Excel-отчёт:
    Лист 1: P&L СВОДНЫЙ
    Лист 2: WB Детализация (по неделям)
    Лист 3: WB Детализация (по месяцам)
    Лист 4: Расходы ИП (из р/с, по категориям)
    Лист 5: FinanceBot (RUB + USDT, по месяцам)
    Лист 6: Самовыкупы
    Лист 7: Налоги (расчёт + сверка)
    Лист 8: Сверка WB ↔ Банк
    Лист 9: Полный журнал

Запуск:
    python -X utf8 dbz_report.py \\
        --bank "Statement.xlsx" \\
        --wb-report "../FinanceSystem/Фин. отчет (Общий).xlsx" \\
        --fb-creds "../FinanceBot/service_account.json" \\
        --fb-sheets "1bFBHIB53h7TJBknJatJn_y__11jCIP45ZbGsatnb1G4" \\
        --buyouts "buyouts_2026.csv" \\
        --period 2026-01 2026-02 \\
        --out "dbz_report_jan-feb_2026.xlsx"

    # Полный 2025 год:
    python -X utf8 dbz_report.py \\
        --bank "Statement.xlsx" \\
        --wb-report "../FinanceSystem/Фин. отчет (Общий).xlsx" \\
        --year 2025 \\
        --out "dbz_report_2025.xlsx"
"""
import argparse
import json
import logging
import sys
from pathlib import Path
from typing import Optional

import pandas as pd

# ─── локальные модули ───────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))

from src.parser import parse_statement                           # noqa: E402
from src.classifier import TransactionClassifier                 # noqa: E402
from src.wb_report import parse_all_sheets, summarize_by_month, net_payout_by_report  # noqa: E402
from src.tax_calc import TaxScheme, calc_tax, extract_tax_payments  # noqa: E402

# fb_adapter требует gspread — импортируем отдельно, чтобы не падать без него
_fb_available = False
try:
    from src.fb_adapter import read_financebot_expenses
    _fb_available = True
except ImportError:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

ENTITY = "DBZ"
OWNER  = "Пирожкова Наталья Викторовна"


# ──────────────────────────────────────────────────────────────────────────
# Вспомогательные функции
# ──────────────────────────────────────────────────────────────────────────

def _load_clusters() -> dict:
    p = Path(__file__).parent / "clusters.json"
    if p.exists():
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _parse_period(period_args: list, year_arg: Optional[int]):
    """Вернуть (date_from, date_to) как pd.Timestamp."""
    if year_arg:
        return pd.Timestamp(f"{year_arg}-01-01"), pd.Timestamp(f"{year_arg}-12-31")
    if period_args and len(period_args) >= 2:
        return pd.Timestamp(period_args[0]), pd.Timestamp(period_args[1]).replace(
            day=pd.Timestamp(period_args[1]).days_in_month
        )
    if period_args and len(period_args) == 1:
        ts = pd.Timestamp(period_args[0])
        return ts, ts.replace(day=ts.days_in_month)
    return None, None


def _filter_period(df: pd.DataFrame, date_from, date_to) -> pd.DataFrame:
    if df.empty or date_from is None:
        return df
    mask = (df["date"] >= date_from) & (df["date"] <= date_to)
    return df[mask].copy()


def _load_buyouts(buyouts_path: Optional[Path]) -> pd.DataFrame:
    """Загрузить данные о самовыкупах из CSV или XLSX."""
    if buyouts_path is None or not buyouts_path.exists():
        logger.info("Самовыкупы: файл не задан или не найден, блок будет пропущен")
        return pd.DataFrame(columns=["date", "amount_gross", "amount_net", "comment"])

    try:
        if buyouts_path.suffix.lower() == ".csv":
            df = pd.read_csv(buyouts_path, parse_dates=["date"])
        else:
            df = pd.read_excel(buyouts_path, parse_dates=["date"])
        logger.info("Самовыкупы: загружено %d строк из '%s'", len(df), buyouts_path.name)
        return df
    except Exception as exc:
        logger.error("Самовыкупы: ошибка чтения '%s': %s", buyouts_path, exc)
        return pd.DataFrame(columns=["date", "amount_gross", "amount_net", "comment"])


def _buyouts_summary(df: pd.DataFrame, date_from, date_to) -> dict:
    if df.empty:
        return {"gross": 0.0, "net": 0.0, "count": 0}
    df = _filter_period(df, date_from, date_to) if date_from else df
    return {
        "gross": df["amount_gross"].sum() if "amount_gross" in df.columns else 0.0,
        "net":   df["amount_net"].sum()   if "amount_net"   in df.columns else 0.0,
        "count": len(df),
    }


# ──────────────────────────────────────────────────────────────────────────
# Расчёт P&L
# ──────────────────────────────────────────────────────────────────────────

def build_pnl(
    wb_df:        pd.DataFrame,
    bank_df:      pd.DataFrame,
    fb_df:        pd.DataFrame,
    buyouts_data: dict,
    date_from,
    date_to,
    year:         int,
    months:       int,
    scheme:       TaxScheme,
) -> pd.DataFrame:
    """
    Строит P&L таблицу в формате строк (метрика, значение, комментарий).

    Структура:
        БЛОК 1: WB Маркетплейс
        БЛОК 2: Расходы ИП (р/с)
        БЛОК 3: FinanceBot (RUB-выплаты)
        БЛОК 3у: FinanceBot USDT (отдельный блок)
        БЛОК 4: Самовыкупы
        БЛОК 5: Налоги (расчётно)
        БЛОК 6: Вывод (информационно)
    """
    rows = []

    def _add(label: str, value, comment: str = "", bold: bool = False, indent: int = 0):
        prefix = "  " * indent
        rows.append({
            "Метрика":    prefix + label,
            "Сумма, руб": round(float(value), 2) if value is not None else None,
            "Комментарий": comment,
            "_bold": bold,
        })

    def _sep(label: str = ""):
        rows.append({"Метрика": label, "Сумма, руб": None, "Комментарий": "", "_bold": True})

    # ── БЛОК 1: WB ───────────────────────────────────────────────────────
    _sep("▌ БЛОК 1: WB МАРКЕТПЛЕЙС")

    def _wb(tx_type: str) -> float:
        if wb_df.empty:
            return 0.0
        return wb_df[wb_df["tx_type"] == tx_type]["amount_rub"].sum()

    sales        = _wb("Продажа WB")
    commission   = _wb("Комиссия WB")      # уже отрицательная
    logistics    = _wb("Логистика WB")     # уже отрицательная
    storage      = _wb("Хранение WB")      # уже отрицательная
    acceptance   = _wb("Приёмка WB")       # уже отрицательная
    fines        = _wb("Штрафы WB")        # уже отрицательная
    other_wb     = _wb("Прочие WB")        # может быть + или −

    # "К перечислению за товар" = Продажи + Комиссия (комиссия отрицательная)
    goods_payout = sales + commission
    # "Итого к получению" = все WB строки вместе
    wb_net = sales + commission + logistics + storage + acceptance + fines + other_wb

    _add("Продажи WB (gross)", sales, "Финотчёт — колонка «Продажа»", bold=True)
    _add("Комиссия WB (с учётом СПП)", commission, "Продажи − К перечислению за товар", indent=1)
    _add("К перечислению за товар", goods_payout, "Финотчёт — «К перечислению за товар»", bold=True)
    _add("Логистика WB", logistics, "Финотчёт — «Стоимость логистики»", indent=1)
    _add("Хранение WB", storage, "Финотчёт — «Стоимость хранения»", indent=1)
    _add("Платная приёмка", acceptance, "Финотчёт — «Стоимость операций на приёмке»", indent=1)
    _add("Штрафы WB", fines, "Финотчёт — «Общая сумма штрафов»", indent=1)
    _add("Реклама + Джем (прочие удержания)", other_wb, "Финотчёт — «Прочие удержания/выплаты»", indent=1)
    _add("ИТОГО К ПОЛУЧЕНИЮ НА Р/С (WB)", wb_net, "Финотчёт — «Итого к оплате»", bold=True)
    _sep()

    # ── БЛОК 2: Расходы ИП из р/с ────────────────────────────────────────
    _sep("▌ БЛОК 2: РАСХОДЫ ИП (р/с МодульБанк)")

    bank_expenses_by_cat: dict = {}
    bank_withdrawal = 0.0

    if not bank_df.empty and "amount_rub" in bank_df.columns:
        expense_rows = bank_df[bank_df["amount_rub"] < 0].copy()

        # Вывод на карты — информационно, не в расходы
        if "tx_type" in expense_rows.columns:
            withdrawals = expense_rows[
                expense_rows["tx_type"].isin(["Перевод (вывод)", "TYPE_TRANSFER_WITHDRAWAL"])
            ]
            bank_withdrawal = abs(withdrawals["amount_rub"].sum())
            expense_rows = expense_rows[
                ~expense_rows["tx_type"].isin(["Перевод (вывод)", "TYPE_TRANSFER_WITHDRAWAL"])
            ]

        if "category" in expense_rows.columns:
            bank_expenses_by_cat = (
                expense_rows.groupby("category")["amount_rub"].sum().to_dict()
            )

    bank_expenses_total = sum(bank_expenses_by_cat.values())

    for cat, amt in sorted(bank_expenses_by_cat.items(), key=lambda x: x[1]):
        _add(cat, amt, indent=1)

    _add("ИТОГО расходы ИП (р/с)", bank_expenses_total, "Из банковской выписки", bold=True)
    _sep()

    # ── БЛОК 3: FinanceBot RUB ────────────────────────────────────────────
    _sep("▌ БЛОК 3: ФИНАНСОВЫЕ ВЫПЛАТЫ (FinanceBot RUB)")

    fb_rub_total = 0.0
    fb_usdt_total = 0.0

    if not fb_df.empty:
        fb_rub  = fb_df[fb_df["currency"] != "USDT"]
        fb_usdt = fb_df[fb_df["currency"] == "USDT"]

        fb_rub_by_cat: dict = {}
        if not fb_rub.empty and "category" in fb_rub.columns:
            fb_rub_by_cat = (
                fb_rub.groupby("category")["amount"].sum().to_dict()
            )
        fb_rub_total = sum(fb_rub_by_cat.values())

        for cat, amt in sorted(fb_rub_by_cat.items(), key=lambda x: -x[1]):
            _add(cat, -abs(amt), indent=1)

        _add("ИТОГО выплаты RUB (FinanceBot)", -fb_rub_total, bold=True)
        _sep()

        # ── БЛОК 3у: FinanceBot USDT ─────────────────────────────────────
        _sep("▌ БЛОК 3у: ВЫПЛАТЫ USDT (отдельный блок)")

        if not fb_usdt.empty:
            fb_usdt_by_cat: dict = {}
            if "category" in fb_usdt.columns:
                fb_usdt_by_cat = fb_usdt.groupby("category")["amount"].sum().to_dict()
            fb_usdt_total = sum(fb_usdt_by_cat.values())

            for cat, amt in sorted(fb_usdt_by_cat.items(), key=lambda x: -x[1]):
                _add(f"{cat} (USDT)", -abs(amt), "в USDT")

            _add("ИТОГО выплаты USDT", -fb_usdt_total, "USDT", bold=True)
            # Справочно в рублях
            usdt_in_rub = fb_usdt["amount_rub"].sum() if "amount_rub" in fb_usdt.columns else 0.0
            _add("USDT ≈ RUB (по курсу)", -abs(usdt_in_rub), "справочно, не суммировать с RUB")
        else:
            _add("Нет USDT-выплат за период", 0.0)

        _sep()

    # ── БЛОК 4: Самовыкупы ───────────────────────────────────────────────
    _sep("▌ БЛОК 4: САМОВЫКУПЫ")

    bo_gross = buyouts_data.get("gross", 0.0)
    bo_net   = buyouts_data.get("net", 0.0)
    bo_count = buyouts_data.get("count", 0)

    _add("Затраты на выкупы (валовые)", -bo_gross, "из файла buyouts")
    _add("Затраты на выкупы (невозвратные)", -bo_net,  "только потери")
    _add(f"Кол-во партий: {bo_count}", None)
    _sep()

    # ── БЛОК 5: Налоги ───────────────────────────────────────────────────
    _sep("▌ БЛОК 5: НАЛОГИ (расчётно)")

    # Доход для УСН = поступления от WB на р/с = wb_net (т.к. это то что пришло на счёт)
    # Точнее — берём из bank_df поступления, помечённые как WB-доход
    usn_income = abs(wb_net)  # приближённо; уточняется по bank_df

    tax_res = calc_tax(
        income_rub=usn_income,
        period_label=f"{date_from.date()} – {date_to.date()}" if date_from else "весь период",
        year=year,
        months_in_period=months,
        scheme=scheme,
    )

    # Фактически уплаченные из р/с
    if not bank_df.empty:
        paid = extract_tax_payments(bank_df)
        tax_res.paid_usn_rub    = paid["paid_usn_rub"]
        tax_res.paid_contrib_rub = paid["paid_contrib_rub"]
        tax_res.paid_ndfl_rub   = paid["paid_ndfl_rub"]

    _add("Доход для УСН (поступления от WB)", usn_income)
    _add("УСН 6% (расчётно)", -tax_res.usn_gross, indent=1)
    _add("Вычет страх. взносов из УСН",  tax_res.usn_deduction, indent=1)
    _add("УСН к уплате (расчётно)",      -tax_res.usn_net,       bold=True)
    _add("1% с превышения 300к",         -tax_res.extra_contrib)
    _add("Взносы ИП фиксированные (за период)", -tax_res.fixed_contrib_period)
    _add("ИТОГО налог + взносы (расчётно)", -tax_res.total_tax, bold=True)
    _add("  Уплачено УСН фактически",    -tax_res.paid_usn_rub,    "из р/с")
    _add("  Уплачено взносов фактически",-tax_res.paid_contrib_rub, "из р/с")
    _add("  Расхождение УСН",             tax_res.diff_usn,
         "расчёт − факт (0 = ок, < 0 = переплата)")
    _sep()

    # ── ИТОГО ────────────────────────────────────────────────────────────
    _sep("▌ ИТОГО")

    operating_profit = (
        wb_net
        + bank_expenses_total
        + (-fb_rub_total)
        + (-bo_net)
        + (-tax_res.total_tax)
    )

    _add("WB NET (к получению на р/с)",      wb_net,            bold=True)
    _add("Расходы ИП (р/с)",                 bank_expenses_total)
    _add("Выплаты FinanceBot RUB",           -fb_rub_total)
    _add("Самовыкупы (невозвратные)",        -bo_net)
    _add("Налоги + взносы (расчётно)",       -tax_res.total_tax)
    _add("ОПЕРАЦИОННАЯ ПРИБЫЛЬ (RUB)",       operating_profit,   bold=True)
    _add("Выплаты USDT (справочно ≈ RUB)",   -abs(usdt_in_rub) if not fb_df.empty and not fb_usdt.empty else 0.0,
         "не включено в прибыль")
    _sep()

    # ── БЛОК 6: Вывод [информационно] ────────────────────────────────────
    _sep("▌ БЛОК 6: ВЫВОД НА ЛИЧНЫЕ КАРТЫ [информационно]")
    _add("Вывод на карты физлица", -bank_withdrawal, "из р/с — внутренний перевод, не расход")
    _sep()

    result_df = pd.DataFrame(rows)
    return result_df


# ──────────────────────────────────────────────────────────────────────────
# Построение листов Excel
# ──────────────────────────────────────────────────────────────────────────

def _style_pnl(ws, df: pd.DataFrame):
    """Применить базовое форматирование к P&L листу."""
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E79")
    section_fill = PatternFill("solid", fgColor="D6E4F0")
    bold_fill    = PatternFill("solid", fgColor="EBF5FB")

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 55

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2  # +1 header, +1 1-based
        metric = str(row.get("Метрика", ""))
        is_bold = bool(row.get("_bold", False))
        is_section = metric.startswith("▌")

        for col_idx in range(1, 4):
            cell = ws.cell(row=excel_row, column=col_idx)
            if is_section:
                cell.fill = section_fill
                cell.font = Font(bold=True, color="1F4E79")
            elif is_bold:
                cell.fill = bold_fill
                cell.font = Font(bold=True)

        # Числовое форматирование колонки B
        b_cell = ws.cell(row=excel_row, column=2)
        if b_cell.value is not None:
            b_cell.number_format = '#,##0'


def export_excel(
    out_path: Path,
    pnl_df:   pd.DataFrame,
    wb_df:    pd.DataFrame,
    bank_df:  pd.DataFrame,
    fb_df:    pd.DataFrame,
    buyouts_df: pd.DataFrame,
    tax_rows: pd.DataFrame,
    wb_reconcile: pd.DataFrame,
):
    """Записать все листы в Excel-файл."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.error("openpyxl не установлен: pip install openpyxl")
        return

    wb = Workbook()
    wb.remove(wb.active)   # удаляем дефолтный лист

    # ── Лист 1: P&L сводный ──────────────────────────────────────────────
    ws1 = wb.create_sheet("P&L Сводный")
    pnl_out = pnl_df.drop(columns=["_bold"], errors="ignore")
    _write_df(ws1, pnl_out)

    # ── Лист 2: WB по неделям ────────────────────────────────────────────
    ws2 = wb.create_sheet("WB по неделям")
    if not wb_df.empty:
        weekly = net_payout_by_report(wb_df)
        _write_df(ws2, weekly)

    # ── Лист 3: WB по месяцам ────────────────────────────────────────────
    ws3 = wb.create_sheet("WB по месяцам")
    if not wb_df.empty:
        monthly = summarize_by_month(wb_df)
        _write_df(ws3, monthly)

    # ── Лист 4: Расходы р/с ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Расходы р-с")
    if not bank_df.empty:
        _write_df(ws4, bank_df.drop(columns=["_bold"], errors="ignore"))

    # ── Лист 5: FinanceBot ────────────────────────────────────────────────
    ws5 = wb.create_sheet("FinanceBot")
    if not fb_df.empty:
        _write_df(ws5, fb_df)

    # ── Лист 6: Самовыкупы ────────────────────────────────────────────────
    ws6 = wb.create_sheet("Самовыкупы")
    if not buyouts_df.empty:
        _write_df(ws6, buyouts_df)

    # ── Лист 7: Налоги ────────────────────────────────────────────────────
    ws7 = wb.create_sheet("Налоги")
    if not tax_rows.empty:
        _write_df(ws7, tax_rows)

    # ── Лист 8: Сверка WB ↔ Банк ─────────────────────────────────────────
    ws8 = wb.create_sheet("Сверка WB-Банк")
    if not wb_reconcile.empty:
        _write_df(ws8, wb_reconcile)

    # ── Лист 9: Полный журнал ─────────────────────────────────────────────
    ws9 = wb.create_sheet("Полный журнал")
    journal = _build_journal(wb_df, bank_df, fb_df)
    if not journal.empty:
        _write_df(ws9, journal)

    wb.save(out_path)
    logger.info("Отчёт сохранён: %s", out_path)


def _write_df(ws, df: pd.DataFrame):
    """Записать DataFrame на лист с заголовком."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    header_fill = PatternFill("solid", fgColor="1F4E79")

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill

    # Авто-ширина (примерная)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)


def _build_journal(wb_df, bank_df, fb_df) -> pd.DataFrame:
    """Собрать единый журнал транзакций из всех источников."""
    frames = []
    cols = ["date", "source", "tx_type", "category", "amount_rub", "amount_usdt",
            "currency", "purpose", "entity"]

    if not wb_df.empty:
        frames.append(wb_df[[c for c in cols if c in wb_df.columns]])

    if not bank_df.empty:
        bank_sub = bank_df.copy()
        if "amount_rub" not in bank_sub.columns and "amount" in bank_sub.columns:
            bank_sub["amount_rub"] = bank_sub["amount"]
        bank_sub["amount_usdt"] = 0.0
        bank_sub["entity"] = ENTITY
        frames.append(bank_sub[[c for c in cols if c in bank_sub.columns]])

    if not fb_df.empty:
        fb_sub = fb_df.copy()
        if "amount_rub" not in fb_sub.columns:
            fb_sub["amount_rub"] = fb_sub["amount_rub"] if "amount_rub" in fb_sub.columns else 0.0
        if "amount_usdt" not in fb_sub.columns:
            fb_sub["amount_usdt"] = 0.0
        frames.append(fb_sub[[c for c in cols if c in fb_sub.columns]])

    if not frames:
        return pd.DataFrame(columns=cols)

    journal = pd.concat(frames, ignore_index=True)
    if "date" in journal.columns:
        journal["date"] = pd.to_datetime(journal["date"])
        journal = journal.sort_values("date").reset_index(drop=True)

    return journal


def _build_wb_reconcile(wb_df: pd.DataFrame, bank_df: pd.DataFrame) -> pd.DataFrame:
    """
    Сверка: NET WB по финотчёту ↔ фактические поступления от WB на р/с.

    NET_wb = SUM("Итого к оплате" из ФинОт за период) — рассчитывается из wb_df
    Bank_wb = поступления на р/с от WB ИНН (из bank_df, категория "Доход WB")
    """
    rows = []

    # WB net по отчётам
    if not wb_df.empty:
        rpt = net_payout_by_report(wb_df)
        for _, r in rpt.iterrows():
            rows.append({
                "Отчёт WB №": r.get("tx_id", ""),
                "Дата":       r.get("date", ""),
                "NET WB (финотчёт)": round(r.get("net_payout_rub", 0), 2),
                "Поступление р/с":   None,
                "Расхождение":       None,
            })

    # Поступления от WB из bank_df
    wb_bank_income = 0.0
    if not bank_df.empty and "category" in bank_df.columns:
        wb_rows = bank_df[
            bank_df["category"].str.lower().str.contains("wb|вайлдберриз|wildberries", na=False)
        ]
        wb_bank_income = wb_rows["amount_rub"].sum() if not wb_rows.empty else 0.0

    if rows:
        total_wb_net = sum(r["NET WB (финотчёт)"] for r in rows)
        rows.append({
            "Отчёт WB №": "ИТОГО",
            "Дата":       "",
            "NET WB (финотчёт)": total_wb_net,
            "Поступление р/с":   wb_bank_income,
            "Расхождение":       round(total_wb_net - wb_bank_income, 2),
        })

    return pd.DataFrame(rows)


# ──────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Полный финансовый отчёт ДБЗ")
    p.add_argument("--bank",       help="МодульБанк р/с XLSX")
    p.add_argument("--wb-report",  help="WB Финотчёт XLSX (Фин. отчет (Общий).xlsx)")
    p.add_argument("--wb-sheet",   default=None, help="Конкретный лист WB (по умолч. все ФинОт ДБЗ*)")
    p.add_argument("--fb-creds",   help="service_account.json для FinanceBot")
    p.add_argument("--fb-sheets",  help="Sheets ID FinanceBot")
    p.add_argument("--buyouts",    help="CSV/XLSX с самовыкупами")
    p.add_argument("--period",     nargs="+", metavar="YYYY-MM",
                   help="Период: --period 2026-01 2026-02 (или один месяц)")
    p.add_argument("--year",       type=int, help="Год целиком: --year 2025")
    p.add_argument("--usn-rate",   type=float, default=6.0, help="Ставка УСН, %% (по умолч. 6)")
    p.add_argument("--out",        default="dbz_report.xlsx", help="Имя выходного файла")
    return p.parse_args()


def main():
    args = parse_args()
    clusters = _load_clusters()

    # ── Период ──────────────────────────────────────────────────────────
    date_from, date_to = _parse_period(args.period, args.year)
    year = args.year or (date_from.year if date_from else pd.Timestamp.now().year)

    if date_from:
        months = (date_to.year - date_from.year) * 12 + date_to.month - date_from.month + 1
    else:
        months = 12

    logger.info("Период: %s → %s (%d мес.)", date_from, date_to, months)

    scheme = TaxScheme(usn_rate=args.usn_rate / 100)

    # ── МодульБанк р/с ──────────────────────────────────────────────────
    bank_df = pd.DataFrame()
    if args.bank:
        bank_path = Path(args.bank)
        if bank_path.exists():
            logger.info("Загружаем банковскую выписку: %s", bank_path.name)
            try:
                raw_bank = parse_statement(bank_path)
                clf = TransactionClassifier()
                classified = [clf.classify(row, owner_name=OWNER) for _, row in raw_bank.iterrows()]
                bank_df = pd.DataFrame(classified)
                # Добавляем подписанную сумму: доход + расход -
                if "amount" in bank_df.columns and "tx_type" in bank_df.columns:
                    bank_df["amount_rub"] = bank_df.apply(
                        lambda r: r["amount"] if r.get("tx_type") in
                        ["Доход", "Возврат", "TYPE_INCOME"] else -abs(r["amount"]),
                        axis=1,
                    )
                if "date" in bank_df.columns:
                    bank_df["date"] = pd.to_datetime(bank_df["date"])
                bank_df = _filter_period(bank_df, date_from, date_to)
                logger.info("Банк р/с: %d операций за период", len(bank_df))
            except Exception as exc:
                logger.error("Ошибка парсинга банка: %s", exc)
        else:
            logger.warning("Файл банка не найден: %s", bank_path)

    # ── WB финотчёт ──────────────────────────────────────────────────────
    wb_df = pd.DataFrame()
    if args.wb_report:
        wb_path = Path(args.wb_report)
        if wb_path.exists():
            logger.info("Загружаем WB финотчёт: %s", wb_path.name)
            if args.wb_sheet:
                from wb_report import parse_wb_excel
                wb_df = parse_wb_excel(wb_path, entity=ENTITY, sheet_name=args.wb_sheet)
            else:
                wb_df = parse_all_sheets(wb_path, entity=ENTITY)
            wb_df = _filter_period(wb_df, date_from, date_to)
            logger.info("WB финотчёт: %d строк за период", len(wb_df))
        else:
            logger.warning("Файл WB финотчёта не найден: %s", wb_path)

    # ── FinanceBot ────────────────────────────────────────────────────────
    fb_df = pd.DataFrame()
    if args.fb_creds and args.fb_sheets and _fb_available:
        logger.info("Подключаемся к FinanceBot Sheets…")
        try:
            entity_patterns = clusters.get("entity_patterns", {})
            fx_rates = clusters.get("fx_rates", {})
            fb_df = read_financebot_expenses(
                sheets_id=args.fb_sheets,
                credentials_path=args.fb_creds,
                entity_patterns=entity_patterns,
                fx_rates=fx_rates,
            )
            # Только DBZ-tagged расходы
            if not fb_df.empty and "entity" in fb_df.columns:
                fb_df = fb_df[fb_df["entity"] == ENTITY]
            fb_df = _filter_period(fb_df, date_from, date_to)
            logger.info("FinanceBot: %d расходов за период", len(fb_df))
        except Exception as exc:
            logger.error("Ошибка FinanceBot: %s", exc)
    elif args.fb_creds and not _fb_available:
        logger.warning("gspread не установлен — FinanceBot пропущен")

    # ── Самовыкупы ────────────────────────────────────────────────────────
    buyouts_path = Path(args.buyouts) if args.buyouts else None
    buyouts_df   = _load_buyouts(buyouts_path)
    buyouts_data = _buyouts_summary(buyouts_df, date_from, date_to)

    # ── P&L ──────────────────────────────────────────────────────────────
    logger.info("Строим P&L…")
    pnl_df = build_pnl(
        wb_df=wb_df,
        bank_df=bank_df,
        fb_df=fb_df,
        buyouts_data=buyouts_data,
        date_from=date_from,
        date_to=date_to,
        year=year,
        months=months,
        scheme=scheme,
    )

    # ── Сверка WB ↔ Банк ─────────────────────────────────────────────────
    wb_reconcile = _build_wb_reconcile(wb_df, bank_df)

    # ── Налоговые строки для отдельного листа ─────────────────────────────
    usn_income = abs(wb_df["amount_rub"].sum()) if not wb_df.empty else 0.0
    tax_res = calc_tax(usn_income, f"{date_from} – {date_to}", year, months, scheme=scheme)
    tax_rows = pd.DataFrame([tax_res.to_dict()])

    # ── Экспорт ──────────────────────────────────────────────────────────
    out_path = Path(args.out)
    logger.info("Экспортируем в %s…", out_path)
    export_excel(
        out_path=out_path,
        pnl_df=pnl_df,
        wb_df=wb_df,
        bank_df=bank_df,
        fb_df=fb_df,
        buyouts_df=buyouts_df,
        tax_rows=tax_rows,
        wb_reconcile=wb_reconcile,
    )

    # ── Краткое резюме в терминал ─────────────────────────────────────────
    print("\n" + "═" * 60)
    print(f"  ОТЧЁТ ДБЗ  |  {date_from.date()} → {date_to.date()}")
    print("═" * 60)
    if not wb_df.empty:
        sales = wb_df[wb_df["tx_type"] == "Продажа WB"]["amount_rub"].sum()
        net   = wb_df["amount_rub"].sum()
        print(f"  WB Продажи:           {sales:>15,.0f} руб")
        print(f"  WB NET (к получению): {net:>15,.0f} руб")
    if not bank_df.empty and "amount_rub" in bank_df.columns:
        exp = bank_df[bank_df["amount_rub"] < 0]["amount_rub"].sum()
        print(f"  Расходы ИП (р/с):     {exp:>15,.0f} руб")
    if not fb_df.empty:
        fb_rub = fb_df[fb_df.get("currency", pd.Series()) != "USDT"]["amount"].sum() if "currency" in fb_df.columns else 0
        print(f"  Выплаты FB RUB:       {-fb_rub:>15,.0f} руб")
    print(f"  Налог расчётный:      {-tax_res.total_tax:>15,.0f} руб")
    print("═" * 60)
    print(f"  Файл: {out_path.absolute()}")
    print()


if __name__ == "__main__":
    main()

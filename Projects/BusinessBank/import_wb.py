#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_wb.py — Импорт WB Финансовых отчётов.

Читает «Финансовый отчёт» (агрегированный) из личного кабинета WB:
    Финансы → Финансовые отчёты → Финансовый отчет → скачать Excel

Использование:
    # Один файл (авто-найдёт все ФинОт ДБЗ-* листы)
    python -X utf8 import_wb.py "../FinanceSystem/Фин. отчет (Общий).xlsx" --entity DBZ

    # С фильтром по периоду
    python -X utf8 import_wb.py report.xlsx --entity DBZ --period 2025-01 2025-12

    # Конкретный лист
    python -X utf8 import_wb.py report.xlsx --entity DBZ --sheet "ФинОт ДБЗ-2 "

    # Несколько файлов / папка
    python -X utf8 import_wb.py file1.xlsx file2.xlsx --entity DBZ
    python -X utf8 import_wb.py --folder "wb_reports/" --entity DBZ

Выходной Excel (3 листа):
    P&L по месяцам  — сводная таблица: месяц × категория затрат + NET
    По неделям      — одна строка на отчёт WB: № отчёта, период, продажи, NET
    Журнал          — полный лог всех транзакций (tx_type + amount_rub)
"""

import argparse
import glob
import io
import logging
import re
import sys
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))

from src.wb_report import (
    net_payout_by_report,
    parse_all_sheets,
    parse_wb_excel,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


# ─── Константы ────────────────────────────────────────────────────────────────

_MONTH_RU = {
    1: "Янв", 2: "Фев", 3: "Мар", 4: "Апр", 5: "Май", 6: "Июн",
    7: "Июл", 8: "Авг", 9: "Сен", 10: "Окт", 11: "Ноя", 12: "Дек",
}

# Порядок отображения типов транзакций в P&L
_TX_TYPES_ORDERED = [
    "Продажа WB",
    "Комиссия WB",
    "Логистика WB",
    "Хранение WB",
    "Приёмка WB",
    "Штрафы WB",
    "Прочие WB",
]

# Колонки журнала для экспорта (остальные UNIFIED_COLUMNS — технические)
_JOURNAL_COLS = ["date", "tx_type", "category", "amount_rub", "purpose", "tx_id"]


# ─── Построение P&L по месяцам ────────────────────────────────────────────────

def build_monthly_pnl(df: pd.DataFrame) -> pd.DataFrame:
    """
    Строит P&L таблицу: одна строка на месяц.

    Columns:
        Период | Продажи | Комиссия | К перечислению | Логистика |
        Хранение | Приёмка | Штрафы | Прочие | NET к получению

    Returns:
        DataFrame с заголовочными строками. Пустой если df пустой.
    """
    if df.empty:
        return pd.DataFrame()

    tmp = df.copy()
    tmp["_month"] = pd.to_datetime(tmp["date"]).dt.to_period("M")

    # Пивот: месяц × tx_type → sum(amount_rub)
    pivot = (
        tmp.groupby(["_month", "tx_type"])["amount_rub"]
        .sum()
        .unstack(fill_value=0.0)
    )

    rows = []
    for period, row in pivot.iterrows():
        ts = period.start_time
        month_label = f"{_MONTH_RU[ts.month]} {ts.year}"

        sales      = row.get("Продажа WB", 0.0)
        commission = row.get("Комиссия WB", 0.0)
        logistics  = row.get("Логистика WB", 0.0)
        storage    = row.get("Хранение WB", 0.0)
        acceptance = row.get("Приёмка WB", 0.0)
        fines      = row.get("Штрафы WB", 0.0)
        other      = row.get("Прочие WB", 0.0)

        # К перечислению за товар = Продажи + Комиссия (комиссия отрицательная)
        k_perechisl = sales + commission
        # NET = все типы вместе = Итого к оплате по финотчёту
        net = sales + commission + logistics + storage + acceptance + fines + other

        rows.append({
            "Период":              month_label,
            "Продажи (gross)":     round(sales, 0),
            "Комиссия WB":         round(commission, 0),
            "К перечислению":      round(k_perechisl, 0),
            "Логистика WB":        round(logistics, 0),
            "Хранение WB":         round(storage, 0),
            "Приёмка WB":          round(acceptance, 0),
            "Штрафы WB":           round(fines, 0),
            "Прочие удержания":    round(other, 0),
            "NET к получению":     round(net, 0),
        })

    result = pd.DataFrame(rows)
    if result.empty:
        return result

    # Строка ИТОГО
    num_cols = [c for c in result.columns if c != "Период"]
    totals = {"Период": "ИТОГО"}
    for col in num_cols:
        totals[col] = result[col].sum()
    result = pd.concat([result, pd.DataFrame([totals])], ignore_index=True)

    return result


# ─── Построение таблицы по неделям ───────────────────────────────────────────

def _extract_period_from_purpose(purpose: str) -> str:
    """Достать 'YYYY-MM-DD – YYYY-MM-DD' из строки purpose."""
    m = re.search(r"\[(.+?)\]", str(purpose))
    return m.group(1) if m else ""


def build_weekly_report(df: pd.DataFrame) -> pd.DataFrame:
    """
    Одна строка на отчёт WB: номер, дата, период, продажи, NET.

    Returns:
        DataFrame с ИТОГО строкой внизу.
    """
    if df.empty:
        return pd.DataFrame()

    weekly = net_payout_by_report(df)
    if weekly.empty:
        return pd.DataFrame()

    # Извлечь период (date_from – date_to) из purpose поля "Продажа WB" строк
    sales_rows = df[df["tx_type"] == "Продажа WB"][["tx_id", "purpose"]].drop_duplicates("tx_id")
    period_map = {
        row["tx_id"]: _extract_period_from_purpose(row["purpose"])
        for _, row in sales_rows.iterrows()
    }

    result = weekly.copy()
    result["Период"] = result["tx_id"].map(period_map).fillna("")
    result["Дата"] = result["date"].dt.strftime("%Y-%m-%d")

    result = result.rename(columns={
        "tx_id":          "Отчёт WB №",
        "net_payout_rub": "NET к получению (руб)",
        "gross_sales":    "Продажи (gross, руб)",
        "rows":           "Строк",
    })

    result = result[[
        "Отчёт WB №", "Дата", "Период",
        "Продажи (gross, руб)", "NET к получению (руб)", "Строк",
    ]]

    # ИТОГО
    totals = {
        "Отчёт WB №":          "ИТОГО",
        "Дата":                 "",
        "Период":               "",
        "Продажи (gross, руб)": result["Продажи (gross, руб)"].sum(),
        "NET к получению (руб)": result["NET к получению (руб)"].sum(),
        "Строк":                result["Строк"].sum(),
    }
    result = pd.concat([result, pd.DataFrame([totals])], ignore_index=True)

    return result


# ─── Excel экспорт ────────────────────────────────────────────────────────────

def _autowidth(ws) -> None:
    """Авто-ширина колонок."""
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)


def _write_sheet(
    writer,
    sheet_name: str,
    df: pd.DataFrame,
    highlight_totals: bool = True,
    num_format: str = "#,##0",
) -> None:
    """Записать DataFrame на лист с форматированием openpyxl."""
    if df.empty:
        pd.DataFrame({"(нет данных)": []}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        return

    df.to_excel(writer, sheet_name=sheet_name, index=False)

    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        ws = writer.sheets[sheet_name]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        totals_fill = PatternFill("solid", fgColor="D6E4F0")
        red_font    = Font(color="C00000")

        # Заголовок
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        max_row = ws.max_row

        # Числовое форматирование + красный для отрицательных
        for row_idx in range(2, max_row + 1):
            for cell in ws[row_idx]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_format
                    if cell.value < 0:
                        cell.font = red_font

        # Последняя строка ИТОГО
        if highlight_totals and max_row > 2:
            for cell in ws[max_row]:
                cell.fill = totals_fill
                if cell.font.color.rgb == "C00000":
                    cell.font = Font(bold=True, color="C00000")
                else:
                    cell.font = Font(bold=True)

        _autowidth(ws)

    except ImportError:
        pass


def export_to_excel(
    pnl: pd.DataFrame,
    weekly: pd.DataFrame,
    journal: pd.DataFrame,
    output_path: Path,
) -> None:
    """Записывает 3 листа в Excel."""
    # Журнал: оставить только полезные колонки
    journal_out = journal[[c for c in _JOURNAL_COLS if c in journal.columns]].copy()
    journal_out = journal_out.rename(columns={
        "date":       "Дата",
        "tx_type":    "Тип",
        "category":   "Категория",
        "amount_rub": "Сумма (руб)",
        "purpose":    "Описание",
        "tx_id":      "Отчёт №",
    })

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_sheet(writer, "P&L по месяцам", pnl)
        _write_sheet(writer, "По неделям", weekly, num_format="#,##0.00")
        _write_sheet(writer, "Журнал", journal_out, highlight_totals=False)

    logger.info("Отчёт сохранён: %s", output_path)


# ─── Консольный вывод ─────────────────────────────────────────────────────────

def print_console_summary(
    pnl: pd.DataFrame,
    weekly: pd.DataFrame,
    journal: pd.DataFrame,
    entity: str,
) -> None:
    """Выводит сводку в консоль."""
    print(f"\n{'='*70}")
    print(f"  WB Финансовый отчёт — {entity}")
    print(f"{'='*70}")

    if not journal.empty:
        date_min = pd.to_datetime(journal["date"]).min().strftime("%Y-%m-%d")
        date_max = pd.to_datetime(journal["date"]).max().strftime("%Y-%m-%d")
        reports_n = journal["tx_id"].nunique()
        print(f"  Период:      {date_min} — {date_max}")
        print(f"  Отчётов WB:  {reports_n}   |   Строк в журнале: {len(journal)}")

    if not pnl.empty:
        totals_row = pnl[pnl["Период"] == "ИТОГО"]
        if not totals_row.empty:
            t = totals_row.iloc[0]
            print()
            print(f"  {'Продажи (gross):':<32} {t.get('Продажи (gross)', 0):>14,.0f} руб")
            print(f"  {'Комиссия WB:':<32} {t.get('Комиссия WB', 0):>14,.0f} руб")
            print(f"  {'К перечислению за товар:':<32} {t.get('К перечислению', 0):>14,.0f} руб")
            print(f"  {'Логистика WB:':<32} {t.get('Логистика WB', 0):>14,.0f} руб")
            print(f"  {'Хранение WB:':<32} {t.get('Хранение WB', 0):>14,.0f} руб")
            print(f"  {'Приёмка WB:':<32} {t.get('Приёмка WB', 0):>14,.0f} руб")
            print(f"  {'Штрафы WB:':<32} {t.get('Штрафы WB', 0):>14,.0f} руб")
            print(f"  {'Прочие удержания:':<32} {t.get('Прочие удержания', 0):>14,.0f} руб")
            print(f"  {'-'*50}")
            print(f"  {'NET к получению на р/с:':<32} {t.get('NET к получению', 0):>14,.0f} руб")

    print(f"{'='*70}")

    if not pnl.empty:
        data_rows = pnl[pnl["Период"] != "ИТОГО"]
        if not data_rows.empty:
            print(f"\n  {'Период':<12} {'Продажи':>13} {'Комиссия':>12} {'Логистика':>12} {'NET':>12}")
            print(f"  {'-'*62}")
            for _, row in data_rows.iterrows():
                print(
                    f"  {row['Период']:<12}"
                    f"{row.get('Продажи (gross)', 0):>13,.0f}"
                    f"{row.get('Комиссия WB', 0):>12,.0f}"
                    f"{row.get('Логистика WB', 0):>12,.0f}"
                    f"{row.get('NET к получению', 0):>12,.0f}"
                )
            print()


# ─── Вспомогательные функции ─────────────────────────────────────────────────

def collect_files(args: argparse.Namespace) -> list[Path]:
    """Дедуплицированный список файлов для обработки."""
    paths: list[Path] = []

    if getattr(args, "folder", None):
        folder = Path(args.folder)
        if not folder.is_dir():
            logger.error("Папка не найдена: %s", folder)
            sys.exit(1)
        for ext in ("*.xlsx", "*.xls"):
            paths.extend(sorted(folder.glob(ext)))

    for pattern in (args.files or []):
        matched = glob.glob(pattern)
        if matched:
            paths.extend(Path(p) for p in sorted(matched))
        else:
            p = Path(pattern)
            if p.exists():
                paths.append(p)
            else:
                logger.warning("Файл не найден: %s", pattern)

    seen: set[Path] = set()
    result: list[Path] = []
    for p in paths:
        key = p.resolve()
        if key not in seen:
            seen.add(key)
            result.append(p)
    return result


def _parse_period(period_args: list | None) -> tuple:
    """Вернуть (date_from, date_to) как pd.Timestamp или (None, None)."""
    if not period_args:
        return None, None
    if len(period_args) >= 2:
        d_from = pd.Timestamp(period_args[0])
        d_to   = pd.Timestamp(period_args[1]).replace(
            day=pd.Timestamp(period_args[1]).days_in_month
        )
        return d_from, d_to
    ts = pd.Timestamp(period_args[0])
    return ts, ts.replace(day=ts.days_in_month)


def _filter_period(df: pd.DataFrame, date_from, date_to) -> pd.DataFrame:
    if df.empty or date_from is None:
        return df
    mask = (df["date"] >= date_from) & (df["date"] <= date_to)
    return df[mask].copy()


# ─── Точка входа ─────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Импорт WB Финансовых отчётов",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "files", nargs="*",
        help="XLSX файлы WB финотчётов (один или несколько)",
    )
    parser.add_argument(
        "--folder", default=None,
        help="Папка — обработать все XLSX внутри",
    )
    parser.add_argument(
        "--entity", default="DBZ",
        help="Код юрлица для фильтрации листов (default: DBZ)",
    )
    parser.add_argument(
        "--sheet", default=None,
        help="Конкретный лист (по умолч. авто-поиск всех ФинОт* листов сущности)",
    )
    parser.add_argument(
        "--period", nargs="+", metavar="YYYY-MM",
        help="Период: --period 2025-01 2025-12  или  --period 2026-01",
    )
    parser.add_argument(
        "--out", default="wb_journal.xlsx",
        help="Имя выходного Excel-файла (default: wb_journal.xlsx)",
    )
    args = parser.parse_args()

    # ── Сбор файлов ──────────────────────────────────────────────────────────
    input_files = collect_files(args)
    if not input_files:
        logger.error("Не найдено ни одного файла. Укажите файлы или --folder.")
        sys.exit(1)

    date_from, date_to = _parse_period(args.period)
    period_label = (
        f"{date_from.date()} — {date_to.date()}" if date_from
        else "весь период"
    )
    logger.info(
        "Файлов: %d | entity: %s | период: %s",
        len(input_files), args.entity, period_label,
    )

    # ── Парсинг ──────────────────────────────────────────────────────────────
    all_dfs: list[pd.DataFrame] = []
    for file_path in input_files:
        try:
            if args.sheet:
                df = parse_wb_excel(file_path, entity=args.entity, sheet_name=args.sheet)
            else:
                df = parse_all_sheets(file_path, entity=args.entity)

            if df.empty:
                logger.warning("Нет данных: %s", file_path.name)
                continue
            all_dfs.append(df)
            logger.info("  ✓ %s — %d строк", file_path.name, len(df))
        except Exception as exc:
            logger.error("  ✗ %s: %s", file_path.name, exc, exc_info=True)

    if not all_dfs:
        logger.error("Ни один файл не обработан успешно.")
        sys.exit(1)

    # ── Объединение и дедупликация ────────────────────────────────────────────
    merged = (
        pd.concat(all_dfs, ignore_index=True)
        .sort_values("date")
        .reset_index(drop=True)
    )

    before = len(merged)
    merged = merged.drop_duplicates(subset=["tx_id", "tx_type"], keep="first")
    dupes = before - len(merged)
    if dupes:
        logger.info("Дедупликация: удалено %d дублей", dupes)

    # ── Фильтр по периоду ────────────────────────────────────────────────────
    if date_from:
        merged = _filter_period(merged, date_from, date_to)
        logger.info("После фильтра по периоду: %d строк", len(merged))

    if merged.empty:
        logger.error("После фильтрации данных нет. Проверьте --period и --entity.")
        sys.exit(1)

    # ── Построение таблиц ────────────────────────────────────────────────────
    pnl    = build_monthly_pnl(merged)
    weekly = build_weekly_report(merged)

    # ── Вывод ────────────────────────────────────────────────────────────────
    print_console_summary(pnl, weekly, merged, args.entity)

    reports_count = merged["tx_id"].nunique()
    print(f"Отчётов WB: {reports_count}   |   Строк журнала: {len(merged)}")

    output_path = Path(args.out)
    export_to_excel(pnl, weekly, merged, output_path)
    print(f"\n✓ Готово! Отчёт: {output_path.resolve()}")


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""Analyze all Excel files - attempt with different modes."""
import sys
import os

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from pathlib import Path

BASE = Path(r"C:\Users\Admin\Documents\workspaces for ai\Cursor\Projects")

# Find xlsx files dynamically
files = []
for p in BASE.iterdir():
    if p.is_dir():
        for f in p.glob("*.xlsx"):
            files.append(f)

files.sort(key=lambda x: x.name)

print(f"Found {len(files)} xlsx files\n")

for fpath in files:
    print("=" * 80)
    print(f"FILE: {fpath.name} ({fpath.stat().st_size / 1024:.1f} KB)")
    print("=" * 80)

    try:
        # Try without read_only to get actual data
        wb = openpyxl.load_workbook(str(fpath), read_only=False, data_only=True)
        print(f"Sheets: {wb.sheetnames}\n")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            print(f"  --- Sheet: '{sheet_name}' ({max_row} rows x {max_col} cols) ---")

            if max_row == 0 or max_col == 0:
                # Try reading cell by cell
                has_data = False
                for r in range(1, 5):
                    for c in range(1, 10):
                        v = ws.cell(row=r, column=c).value
                        if v is not None:
                            has_data = True
                            print(f"    [{r},{c}] = {str(v)[:80]}")
                if not has_data:
                    print("    (empty sheet)")
                print()
                continue

            # Print headers (row 1-3) and sample data
            for r in range(1, min(8, max_row + 1)):
                vals = []
                for c in range(1, min(25, max_col + 1)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        col_letter = openpyxl.utils.get_column_letter(c)
                        vals.append(f"{col_letter}={str(v)[:50]}")
                if vals:
                    print(f"    Row {r}: {' | '.join(vals)}")

            if max_row > 8:
                print(f"    ... ({max_row - 8} more rows) ...")
                # Show last 2 rows
                for r in range(max(9, max_row - 1), max_row + 1):
                    vals = []
                    for c in range(1, min(25, max_col + 1)):
                        v = ws.cell(row=r, column=c).value
                        if v is not None:
                            col_letter = openpyxl.utils.get_column_letter(c)
                            vals.append(f"{col_letter}={str(v)[:50]}")
                    if vals:
                        print(f"    Row {r}: {' | '.join(vals)}")
            print()

        wb.close()
    except Exception as e:
        print(f"  ERROR: {type(e).__name__}: {e}\n")

print("\n[DONE]")

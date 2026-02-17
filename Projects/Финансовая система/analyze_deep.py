# -*- coding: utf-8 -*-
"""Deep analysis of specific sheets."""
import sys
import os

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from pathlib import Path

BASE = Path(r"C:\Users\Admin\Documents\workspaces for ai\Cursor\Projects")

# Find files dynamically
files = {}
for p in BASE.iterdir():
    if p.is_dir():
        for f in p.glob("*.xlsx"):
            files[f.name] = f

# 1. Read "файл с описанием элементов системы" fully
print("=" * 80)
print("DEEP READ: файл с описанием элементов системы.xlsx")
print("=" * 80)

for name, fpath in files.items():
    if "описани" in name.lower():
        wb = openpyxl.load_workbook(str(fpath), read_only=False, data_only=True)
        ws = wb['Элементы']
        for r in range(1, min(30, ws.max_row + 1)):
            vals = []
            for c in range(1, min(27, ws.max_column + 1)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    vals.append(f"{col_letter}={str(v)[:120]}")
            if vals:
                print(f"  Row {r}: {' | '.join(vals)}")
        wb.close()

# 2. Read full Баланс sheet from МН journal
print("\n" + "=" * 80)
print("DEEP READ: Журнал МН - Баланс (all rows)")
print("=" * 80)

for name, fpath in files.items():
    if "МН" in name:
        wb = openpyxl.load_workbook(str(fpath), read_only=False, data_only=True)
        ws = wb['Баланс']
        for r in range(1, ws.max_row + 1):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    vals.append(f"{col_letter}={str(v)[:80]}")
            if vals:
                print(f"  Row {r}: {' | '.join(vals)}")
        wb.close()

# 3. Read Баланс счетов - all rows from Февраль
print("\n" + "=" * 80)
print("DEEP READ: Баланс счетов - Февраль 2026 (all rows)")
print("=" * 80)

for name, fpath in files.items():
    if "Баланс счетов" in name:
        wb = openpyxl.load_workbook(str(fpath), read_only=False, data_only=True)
        for sn in wb.sheetnames:
            if "Февраль" in sn:
                ws = wb[sn]
                for r in range(1, ws.max_row + 1):
                    vals = []
                    for c in range(1, min(15, ws.max_column + 1)):
                        v = ws.cell(row=r, column=c).value
                        if v is not None:
                            col_letter = openpyxl.utils.get_column_letter(c)
                            vals.append(f"{col_letter}={str(v)[:80]}")
                    if vals:
                        print(f"  Row {r}: {' | '.join(vals)}")
        wb.close()

# 4. Пользователи - all rows
print("\n" + "=" * 80)
print("DEEP READ: Выплаты - Пользователи (all rows)")
print("=" * 80)

for name, fpath in files.items():
    if "Выплаты" in name:
        wb = openpyxl.load_workbook(str(fpath), read_only=False, data_only=True)
        ws = wb['Пользователи']
        for r in range(1, ws.max_row + 1):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    vals.append(f"{col_letter}={str(v)[:80]}")
            if vals:
                print(f"  Row {r}: {' | '.join(vals)}")
        wb.close()

# 5. Крипто - Расчетный баланс (all rows)
print("\n" + "=" * 80)
print("DEEP READ: Крипто - Расчетный баланс (all rows)")
print("=" * 80)

for name, fpath in files.items():
    if "Крипто" in name:
        wb = openpyxl.load_workbook(str(fpath), read_only=False, data_only=True)
        ws = wb['Расчетный баланс']
        for r in range(1, ws.max_row + 1):
            vals = []
            for c in range(1, min(10, ws.max_column + 1)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    vals.append(f"{col_letter}={str(v)[:100]}")
            if vals:
                print(f"  Row {r}: {' | '.join(vals)}")
        wb.close()

# 6. Справочник from МН
print("\n" + "=" * 80)
print("DEEP READ: Журнал МН - Справочник (first 50 rows)")
print("=" * 80)

for name, fpath in files.items():
    if "МН" in name:
        wb = openpyxl.load_workbook(str(fpath), read_only=False, data_only=True)
        ws = wb['Справочник']
        for r in range(1, min(50, ws.max_row + 1)):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    vals.append(f"{col_letter}={str(v)[:80]}")
            if vals:
                print(f"  Row {r}: {' | '.join(vals)}")
        wb.close()

print("\n[DONE]")

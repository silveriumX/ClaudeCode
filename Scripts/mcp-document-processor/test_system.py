"""
Автоматический тест MCP Document Processor
Проверяет: API ключ, библиотеки, обработку документов
"""

import os
import sys
from pathlib import Path

print("=" * 60)
print("АВТОМАТИЧЕСКИЙ ТЕСТ MCP DOCUMENT PROCESSOR")
print("=" * 60)
print()

# 1. Проверка API ключа
print("[1/5] Проверка API ключа...")
api_key = os.getenv("ANTHROPIC_API_KEY")
if not api_key:
    print("[FAIL] ANTHROPIC_API_KEY not set in .env")
    sys.exit(1)
print("[OK] API ключ установлен")
print()

# 2. Проверка библиотек
print("[2/5] Проверка установленных библиотек...")
try:
    import anthropic
    print(f"[OK] anthropic {anthropic.__version__}")
except ImportError as e:
    print(f"[ERR] Ошибка: {e}")
    sys.exit(1)

try:
    from openpyxl import Workbook
    print("[OK] openpyxl установлен")
except ImportError as e:
    print(f"[ERR] Ошибка: {e}")
    sys.exit(1)

try:
    from PIL import Image
    print("[OK] Pillow установлен")
except ImportError as e:
    print(f"[ERR] Ошибка: {e}")
    sys.exit(1)

print()

# 3. Создание тестового Excel файла
print("[3/5] Создание тестового Excel файла...")
test_file = Path(__file__).parent / "test_document.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Тестовые данные"

# Добавляем данные
ws['A1'] = "Продукт"
ws['B1'] = "Количество"
ws['C1'] = "Цена"

ws['A2'] = "Ноутбук"
ws['B2'] = 5
ws['C2'] = 50000

ws['A3'] = "Мышь"
ws['B3'] = 20
ws['C3'] = 500

ws['A4'] = "Клавиатура"
ws['B4'] = 15
ws['C4'] = 1500

wb.save(test_file)
print(f"[OK] Файл создан: {test_file}")
print()

# 4. Обработка через Claude API
print("[4/5] Отправка файла в Claude API...")
print("Модель: claude-sonnet-4-5")
print()

from openpyxl import load_workbook

workbook = load_workbook(test_file, data_only=True)
result_parts = []

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    result_parts.append(f"## {sheet_name}\n")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue

    max_cols = max(len(row) for row in rows if row)

    for i, row in enumerate(rows):
        if not any(row):
            continue

        row_data = [str(cell) if cell is not None else "" for cell in row[:max_cols]]
        result_parts.append("| " + " | ".join(row_data) + " |")

        if i == 0:
            result_parts.append("| " + " | ".join(["---"] * len(row_data)) + " |")

excel_content = "\n".join(result_parts)

# Отправляем в Claude
client = anthropic.Anthropic(api_key=api_key)

response = client.messages.create(
    model="claude-sonnet-4-5",
    max_tokens=2000,
    messages=[{
        "role": "user",
        "content": f"Проанализируй эти данные из Excel и сделай краткое резюме:\n\n{excel_content}"
    }]
)

result = response.content[0].text

print("[OK] Ответ получен от Claude!")
print()

# 5. Результат
print("[5/5] РЕЗУЛЬТАТ ОБРАБОТКИ:")
print("-" * 60)

# Сохраняем в файл чтобы избежать проблем с кодировкой
result_file = Path(__file__).parent / "test_result.txt"
with open(result_file, "w", encoding="utf-8") as f:
    f.write(result)
    f.write("\n\n")
    f.write("=" * 60 + "\n")
    f.write("СТАТИСТИКА:\n")
    f.write(f"- Входных токенов: {response.usage.input_tokens}\n")
    f.write(f"- Выходных токенов: {response.usage.output_tokens}\n")
    f.write(f"- Стоимость: ~${(response.usage.input_tokens * 3 + response.usage.output_tokens * 15) / 1_000_000:.4f}\n")

print(f"Результат сохранен в: {result_file}")
print("-" * 60)
print()

print("=" * 60)
print("[SUCCESS] ТЕСТ ПРОЙДЕН УСПЕШНО!")
print("=" * 60)
print()
print("Система готова к работе!")
print("После перезапуска Cursor можно использовать:")
print('  "Обработай файл C:\\path\\to\\file.pdf"')
